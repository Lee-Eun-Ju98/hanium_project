import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime



url_format = 'https://search.naver.com/search.naver?where=news&query={0}&sm=tab_opt&sort=1&photo=0&field=0&pd=3&ds={1}&de={2}&docid=&related=0&mynews=1&office_type=1&office_section_code=3&news_office_checked=1009&nso=so%3Add%2Cp%3Afrom{3}to{4},a:all&start={5}'

searching = input('검색어를 입력하세요 : ')
number = int(input('몇개의 기사를 가져올까요??(10단위) : '))
firstdate = input('검색할 날짜 시작일 ex) 2021.01.01 : ')
lastdate = input('검색할 날짜 마지막날 ex) 2021.07.31 : ')


firstdate2 = datetime.datetime.strptime(firstdate, '%Y.%m.%d')
lastdate2 = datetime.datetime.strptime(lastdate, '%Y.%m.%d')

ds = firstdate2.strftime('%Y.%m.%d')
de = firstdate2.strftime('%Y.%m.%d')

wb = openpyxl.Workbook()
sheet = wb.active

count = 2
articles = []
article = 0
dates = ''
news_r = []
# 페이지 크롤링
for page_number in range(number):
    r = requests.get(url_format.format(searching, firstdate,
                                       lastdate, ds, de, str(page_number*10+1)))
    sp = BeautifulSoup(r.text, 'html.parser')
    sources = sp.select('div.group_news > ul.list_news > li div.news_area > a')

    # 뉴스 기사 크롤링
    for source in sources:
        news_r = requests.get(source.attrs['href'], headers={
                              'User-Agent': 'Mozilla/5.0'})
        news_sp = BeautifulSoup(news_r.content.decode(
            'euc-kr', 'replace'), 'html.parser')
        article = news_sp.select('#article_body')
        dates = news_sp.select(
            '#top_header > div > div > div.news_title_author > ul > li.lasttime')
        try:
            sheet.cell(row=count, column=1).value = dates[0].text
            sheet.cell(row=count, column=2).value = article[0].text
            count += 1
        except:
            print("오류 발생")

sheet.cell(row=1, column=1).value = str(searching) + ' 크롤링 결과'
sheet.cell(row=1, column=2).value = datetime.datetime.now()


save = searching + '.xlsx'

wb.save(save)
wb.close()
